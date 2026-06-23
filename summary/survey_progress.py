import os
os.chdir('../')

import json
import yaml
import joblib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from adjustText import adjust_text
from astropy.io import fits
from astropy.visualization import make_lupton_rgb # Creating color RGB images
            
from data_prep.image.read_and_cutout import read_subaru_img_wcs, cutoutimg
from data_prep.spec.read_and_cutout  import readinfodat
from core.fitting_result_utils import complete_fit_params
from core.plot_corner  import read_post
from diagnostics.bad_g1g2_identifier import check_g1g2_post

from klm.safe_plot import setup; setup() # must before plt
plt.style.use('default')
plt.rcParams.update({
    "font.family": "sans-serif",
    "font.sans-serif": "Helvetica",
    "font.serif": "Helvetica",
})



def posterior_to_cat(g_cat, post_dir, ztable, slits_not_finished):
    """
        Generate a shear catalog.
    """
    for slit_num in range(1, 143):
        z_spec = ztable[slit_num, 1]
        
        # Match RA/DEC
        infdatfilename = f'info.829.{slit_num:03d}.{slit_num+100305:06d}.dat'
        dat_dict       = readinfodat(spec1d2dfolder02 + infdatfilename)
        g_cat[slit_num+1, 15:17] = dat_dict['RA'], dat_dict['DEC']
        
        if np.isnan(z_spec)==True:
            print( "\033[43m" + 'WARNING: ' + "\033[0m " + 
                  f'Slit {slit_num} skipped because no secure redshift found.\n')
            continue
    
        elif np.isnan(z_spec)==False:
            # Load pkl
            try:
                with open(f'{real_pkl_folder}pkl/slit_{slit_num:03d}.pkl', "rb") as f:
                    data_info = joblib.load(f)
            except FileNotFoundError: 
                print( "\033[43m" + 'WARNING: ' + "\033[0m " + 
                      f'Slit {slit_num} skipped because no pkl found.\n')
                continue
            
            # Check if already solved M_stellar
            if data_info['galaxy']['log10_Mstar'] is None:
                # Read M_stellar
                dg = pd.read_csv(f"{Ms_folder}Mstellar_table.txt", 
                                 sep=r"\s+", header=0, 
                                 names=["slit","median","std","err_lo","mean","err_hi"])
                dg = dg.drop_duplicates(subset="slit", keep="last")
                Ms = dg.sort_values(by='slit').to_numpy()
                for row in Ms: 
                    if row[0]==slit_num: 
                        log10_Mstar = row[1]
                        log10_Mstar_err = row[2]
                        break
            else:
                log10_Mstar = data_info['galaxy']['log10_Mstar']
                log10_Mstar_err = data_info['galaxy']['log10_Mstar_err']
            g_cat[slit_num+1, 17:19] = log10_Mstar, log10_Mstar_err
            del log10_Mstar, log10_Mstar_err
            
            if slit_num in slits_not_finished:
                print( "\033[43m" + 'WARNING: ' + "\033[0m " + 
                      f'Slit {slit_num} skipped since you specified the fitting is not finished.\n')
                continue
            
            full_run_dir_1 = f'{post_dir}Slit_{slit_num:03d}/'
            if Path(full_run_dir_1).exists() is False:
                print(f'Slit {slit_num} does not exist.\n')
                continue
            
            post_path1 = f'{full_run_dir_1}post.txt'
            best_path1 = f'{full_run_dir_1}best_fit.json'
            
            run_samples1, par_names1, weights1, \
            loglikes1, samples1, mask1, no_plot1 = read_post(
                post_path1, percentile=95)
            # par_names1 = run_samples1[0,  2:]
            samples1   = run_samples1[1:, 2:].astype(float)
            
            # Posterior txt --> dict
            alllinespecies = []
            for spec in data_info['spec']:
                alllinespecies.append(spec['meta']['line_species'])
            percentile = analyze_percentile(samples1, 
                                            "config/binospec_fitting_params.yaml", 
                                            alllinespecies)
            
            if not isinstance(percentile, dict): 
                print( "\033[43m" + 'WARNING:' + "\033[0m " + 
                      f'Slit {slit_num} hint: your run does not match your current pkl or config yaml. Skipped.\n')
                continue
            
            # Add best params into percentile array
            with open(best_path1, "r") as f:
                best_par = json.load(f)['maximum_likelihood']['point']
            for key, vals in percentile.items():
                percentile[key]['best'] = best_par[key]
            
            # Add more columns to contain posteriors
            n_rows, n_cols = g_cat.shape
            for key in list(percentile.keys()):
                # We only do once in the slit for loop!
                if n_cols <= 19: 
                    if key.split('-')[0] == 'shared_params':
                        par = key.split('-')[1]
                        g_cat = np.append(
                            g_cat, 
                            np.array([
                                [np.nan, f'{par}']+[np.nan]*(n_rows-2), # col 19
                                [np.nan, f'{par}_median']+[np.nan]*(n_rows-2), # col 20
                                [np.nan, f'{par}_err-']+[np.nan]*(n_rows-2), # col 21
                                [np.nan, f'{par}_err+']+[np.nan]*(n_rows-2), # col 22
                                ]).T, 
                            axis=1)
                
                # Assign values from posterior
                par = key.split('-')[1]
                for j in range(len(g_cat[1])):
                    head_key = g_cat[1, j]
                    if par == head_key:
                        g_cat[slit_num+1, j:j+4] = list([
                            percentile[key]['best'], 
                            percentile[key]['median'], # or mean
                            percentile[key]['err_lo'], 
                            percentile[key]['err_hi'], 
                            ])
    
            print("\033[42m" + 'INFO:    ' + "\033[0m " + 
                  f'Slit {slit_num} recorded. 👍\n')
    
    return g_cat


def analyze_percentile(samples, config_filename, 
                       line_species=['O2','O2','O2','Hg','Hg','Hg']):
    # load fitting config
    with open(config_filename, "r", encoding="utf-8") as file1:
        config = yaml.safe_load(file1)
    
    config_par = complete_fit_params(config, line_species)
    
    config_param_names = [key for key, _ in config_par.items()]
    
    dic_percent = {}
    nparams = len(samples[0])
    arr = np.zeros((nparams, 3))
    
    for j in range(nparams):
        samp_points = samples[:,j]
        x123 = np.percentile(samp_points, [16, 50, 84])
        err_lo, median, err_hi = x123[0]-x123[1], x123[1], x123[2]-x123[1]
        
        if ('-g1' in config_param_names[j]) or \
            ('-g2' in config_param_names[j]):
                
            is_g_bad = check_g1g2_post(
                samp_points, pmin=-0.15, pmax=0.15
                )
            
            if is_g_bad:
                # flip signs of low/high errors on purpose
                err_lo *= -1
                err_hi *= -1
                print(f'Bad shear detected: {config_param_names[j]}')
        
        arr[j] = np.around([median, err_lo, err_hi], decimals=4)
        
        try:
            dic_percent[config_param_names[j]] = {'median': arr[j,0], 
                                                  'err_lo': arr[j,1], 
                                                  'err_hi': arr[j,2]}
        except IndexError:
            dic_percent = 'ERROR'
            break

    return dic_percent


def binospec_fov(ax, ctr_pt,rotatn):
    ra0, de0 = ctr_pt[0], ctr_pt[1]
    cnrA1r = ra0 + ((-480-96)/3600)*np.cos(rotatn) - ((+450)/3600)*np.sin(rotatn)
    cnrA1d = de0 + ((-480-96)/3600)*np.sin(rotatn) + ((+450)/3600)*np.cos(rotatn)
    cnrA2r = ra0 + ((-000-96)/3600)*np.cos(rotatn) - ((+450)/3600)*np.sin(rotatn)
    cnrA2d = de0 + ((-000-96)/3600)*np.sin(rotatn) + ((+450)/3600)*np.cos(rotatn)
    cnrA3r = ra0 + ((-000-96)/3600)*np.cos(rotatn) - ((-450)/3600)*np.sin(rotatn)
    cnrA3d = de0 + ((-000-96)/3600)*np.sin(rotatn) + ((-450)/3600)*np.cos(rotatn)
    cnrA4r = ra0 + ((-480-96)/3600)*np.cos(rotatn) - ((-450)/3600)*np.sin(rotatn)
    cnrA4d = de0 + ((-480-96)/3600)*np.sin(rotatn) + ((-450)/3600)*np.cos(rotatn)
    Ax_values = [cnrA1r, cnrA2r, cnrA3r, cnrA4r, cnrA1r]
    Ay_values = [cnrA1d, cnrA2d, cnrA3d, cnrA4d, cnrA1d]
    cnrB1r = ra0 + ((+480+96)/3600)*np.cos(rotatn) - ((+450)/3600)*np.sin(rotatn)
    cnrB1d = de0 + ((+480+96)/3600)*np.sin(rotatn) + ((+450)/3600)*np.cos(rotatn)
    cnrB2r = ra0 + ((+000+96)/3600)*np.cos(rotatn) - ((+450)/3600)*np.sin(rotatn)
    cnrB2d = de0 + ((+000+96)/3600)*np.sin(rotatn) + ((+450)/3600)*np.cos(rotatn)
    cnrB3r = ra0 + ((+000+96)/3600)*np.cos(rotatn) - ((-450)/3600)*np.sin(rotatn)
    cnrB3d = de0 + ((+000+96)/3600)*np.sin(rotatn) + ((-450)/3600)*np.cos(rotatn)
    cnrB4r = ra0 + ((+480+96)/3600)*np.cos(rotatn) - ((-450)/3600)*np.sin(rotatn)
    cnrB4d = de0 + ((+480+96)/3600)*np.sin(rotatn) + ((-450)/3600)*np.cos(rotatn)
    Bx_values = [cnrB1r, cnrB2r, cnrB3r, cnrB4r, cnrB1r]
    By_values = [cnrB1d, cnrB2d, cnrB3d, cnrB4d, cnrB1d]
    ax.plot(Ax_values, Ay_values, color='#2176ff', # BinoSpec FoV-1A
            transform=ax.get_transform('world'), zorder=2)
    ax.plot(Bx_values, By_values, color='#2176ff', # BinoSpec FoV-1B
            transform=ax.get_transform('world'), zorder=2)
    return 


def scale_calc(z, H0, Omega_M, Omega_Lam):
    from astropy.cosmology import FlatLambdaCDM
    cosmo = FlatLambdaCDM(H0, Om0=Omega_M) # a selected cosmological model
    d_A = cosmo.kpc_proper_per_arcmin(z)
    d_A = d_A.value/60 # kpc/arcmin --> kpc/arcsec
    return d_A


def draw_r500_r200(plt, ax, r_500, r_200, color='greenyellow'):
    circ_500 = plt.Circle(( (( 2)+(48)/60+( 3.3)/3600)*15,
                           -(( 3)+(31)/60+(46.4)/3600)   ), 
                          r_500/3600, 
                          color=color, fill=False, linestyle='--', # 'dashdot'
                          linewidth=1.2, zorder=1, 
                          transform=ax.get_transform('world'))
    circ_200 = plt.Circle(( (( 2)+(48)/60+( 3.3)/3600)*15,
                           -(( 3)+(31)/60+(46.4)/3600)   ), 
                          r_200/3600, 
                          color=color, fill=False, linestyle=':', 
                          zorder=1, 
                          transform=ax.get_transform('world'))
    ax.text( (( 2)+(48)/60+( 3.3)/3600)*15,
             -(( 3)+(31)/60+(46.4)/3600) - r_500/3600,
             r'$r_{\rm 500}$', color=color, size=18, 
             verticalalignment='top', 
             transform=ax.get_transform('world'))
    ax.text( (( 2)+(48)/60+( 3.3)/3600)*15,
             -(( 3)+(31)/60+(46.4)/3600) - r_200/3600,
             r'$r_{\rm 200}$', color=color, size=18, 
             verticalalignment='top', 
             transform=ax.get_transform('world'))
    ax.add_patch(circ_500)
    ax.add_patch(circ_200)
    return circ_500, circ_200


def slit_design_plot(g_cat, maskCenterRA, maskCenterDEC, imgR, imgG, imgB, wcs, 
                     slits_not_finished=None):
    assert ~np.any(np.isnan(g_cat[2:, 15:17].astype(float))), \
        'Some slits have no RA/DEC. Maybe slits were not iterated completed?'
    
    plt.style.use('dark_background')
    
    fig = plt.figure(figsize=(8,8), dpi=450)
    plt.subplots_adjust(hspace=0, wspace=0) # h=height
    gs = fig.add_gridspec(1, 1,
                          height_ratios=[1],
                          width_ratios=[1])
    ax1 = fig.add_subplot(gs[0, 0], projection=wcs)
    
    g_cat_return = {}
    
    # (1) Slits without secure z and M*
    z_arr = np.array(g_cat[2:,  1], dtype=float) # Must use float
    Ms_ar = np.array(g_cat[2:, 17], dtype=float) # Must use float
    no_z  = np.isnan(z_arr)
    no_Ms = np.isnan(Ms_ar)
    have_z_and_Ms = ~no_z & ~no_Ms
    g_cat_secure = g_cat[np.r_[True, True,  have_z_and_Ms]]
    g_cat_no_z   = g_cat[np.r_[True, True, ~have_z_and_Ms]]
    g_cat_return['g_cat_no_z'] = g_cat_no_z
    print(f'No secure z ({len(g_cat_no_z[2:, 0])}): {g_cat_no_z[2:, 0]}')
    
    # (2) Slits with z but not yet solve posterior
    g1_arr = np.array(g_cat_secure[2:, 19], dtype=float)
    tbd    = np.isnan(g1_arr)
    if slits_not_finished is not None:
        tbd = np.isnan(g1_arr) | np.isin(g_cat_secure[2:, 0], np.array(slits_not_finished))
    g_cat_solved = g_cat_secure[np.r_[True, True, ~tbd]]
    g_cat_tbd    = g_cat_secure[np.r_[True, True,  tbd]]
    g_cat_return['g_cat_tbd'] = g_cat_tbd
    print(f'Not finished ({len(g_cat_tbd[2:, 0])}): {g_cat_tbd[2:, 0]}')
    
    # (3) Slits with z & posterior solved but not valid (FAILED)
    g1_errs = np.array(g_cat_solved[2:, 20], dtype=float)
    g_cat_fail  = g_cat_solved[np.r_[True, True,  np.isnan(g1_errs)]]
    g_cat_valid = g_cat_solved[np.r_[True, True, ~np.isnan(g1_errs)]]
    g_cat_return['g_cat_fail'] = g_cat_fail
    print(f'Failed ({len(g_cat_fail[2:, 0])}): {g_cat_fail[2:, 0]}')
    
    # (4) Slits with z & posterior solved & valid but extreme g1/g2
    g1err1 = np.array(g_cat_valid[2:, 21], dtype=float) # should be -
    g1err2 = np.array(g_cat_valid[2:, 22], dtype=float) # should be +
    g2err1 = np.array(g_cat_valid[2:, 25], dtype=float) # should be -
    g2err2 = np.array(g_cat_valid[2:, 26], dtype=float) # should be +
    
    mask_good_g1 = (g1err1 < 0) & (g1err2 > 0)
    mask_good_g2 = (g2err1 < 0) & (g2err2 > 0)
    mask_good_g1g2  = mask_good_g1 & mask_good_g2
    g_cat_good_g1g2 = g_cat_valid[np.r_[True, True,  mask_good_g1g2]]
    g_cat_bad_g1g2  = g_cat_valid[np.r_[True, True, ~mask_good_g1g2]]
    g_cat_return['g_cat_good_g1g2'] = g_cat_good_g1g2
    g_cat_return['g_cat_bad_g1g2' ] = g_cat_bad_g1g2
    print(f'Bad g1g2 ({len(g_cat_bad_g1g2[2:, 0])}): {g_cat_bad_g1g2[2:, 0]}')
    print(f'Good g1g2 ({len(g_cat_good_g1g2[2:, 0])}): {g_cat_good_g1g2[2:, 0]}')
    
    ax1.scatter(g_cat_no_z[2:, 15], 
                g_cat_no_z[2:, 16], 
                label=f'No redshift or M* ({len(g_cat_no_z[2:, 15])})', 
                marker='x',
                s=30, color='gray', 
                transform=ax1.get_transform('world'), zorder=4)
    if len(g_cat_tbd[2:, 15]) != 0:
        ax1.scatter(g_cat_tbd[2:, 15], 
                    g_cat_tbd[2:, 16], 
                    label=f'Running on HPC ({len(g_cat_tbd[2:, 15])})', 
                    marker='o',
                    s=30, facecolors='gray', edgecolors='black', 
                    transform=ax1.get_transform('world'), zorder=4)
    ax1.scatter(g_cat_bad_g1g2[2:, 15], 
                g_cat_bad_g1g2[2:, 16], 
                label=r'$g_1$'+' or '+r'$g_2$'+' rejected '+
                f'({len(g_cat_bad_g1g2[2:, 15])})', 
                marker='^',
                s=40, facecolors='violet', edgecolors='black', 
                transform=ax1.get_transform('world'), zorder=4)
    ax1.scatter(g_cat_good_g1g2[2:, 15], 
                g_cat_good_g1g2[2:, 16], 
                label=r'$g_1$'+' and '+r'$g_2$'+' accepted '+
                f'({len(g_cat_good_g1g2[2:, 15])})', 
                marker='o',
                s=40, facecolors='cyan', edgecolors='black', # '#0000ff'
                transform=ax1.get_transform('world'), zorder=4)
    
    # Annotate slits & Avoid overlapping
    texts = []
    # Not reliable fit
    for x, y, slit_id in zip(g_cat_fail[2:, 15]-3/3600, #  x
                             g_cat_fail[2:, 16]+3/3600, #  y
                             g_cat_fail[2:,  0]): # slit ID
        texts.append(
            ax1.text(x, y, str(slit_id), color='red', 
                     transform=ax1.get_transform('world'), fontsize=8, zorder=3)
                     )
    # Extreme g1 or g2
    for x, y, slit_id in zip(g_cat_bad_g1g2[2:, 15]-3/3600, #  x
                             g_cat_bad_g1g2[2:, 16]+3/3600, #  y
                             g_cat_bad_g1g2[2:,  0]): # slit ID
        texts.append(
            ax1.text(x, y, str(slit_id), color='mediumorchid', 
                     transform=ax1.get_transform('world'), fontsize=6, zorder=4)
                     )
    # Solved
    for x, y, slit_id in zip(g_cat_good_g1g2[2:, 15]-3/3600, #  x
                             g_cat_good_g1g2[2:, 16]+3/3600, #  y
                             g_cat_good_g1g2[2:,  0]): # slit ID
        texts.append(
            ax1.text(x, y, str(slit_id), color='cyan', # '#0000ff'
                     transform=ax1.get_transform('world'), fontsize=8, zorder=5)
                     )
    adjust_text(
        texts,
        ax=ax1,
        expand_points=(1.2, 1.2),
        expand_text=(1.2, 1.2),
        arrowprops=dict(
            arrowstyle='-',
            lw=0.5,
            shrinkA=5, 
            shrinkB=5), 
        verbose=True
    )

    # BinoSpec FoV
    maskCenter = [maskCenterRA, maskCenterDEC]
    maskPA_rad = -25 / 57.3 # left-->west right-->east
    binospec_fov(ax1, maskCenter, maskPA_rad)
    
    # R_500, R_200(=R_500/0.65)
    R_500 = 944 # kpc, Vikhlinin et al. (2006) arXiv:astro-ph/0507092, Table 4
    d_A   = scale_calc(0.1883, 70, 0.3, 0.7) # kpc/arcsec
    r_500 = R_500 / d_A
    r_200 = r_500 / 0.65
    c500, c200 = draw_r500_r200(plt, ax1, r_500, r_200)
    
    # 1-arcsec-scale
    ax1.arrow(x=(41.95), y=(-3.66), dx=-(5)/60, dy=0, 
              head_width=0, head_length=0, 
              fc='white', ec='white', width=10/3600, 
              transform=ax1.get_transform('world'))
    ax1.text((41.95)-(2.5)/60, (-3.66), 
             '5'+'\'', size=18,
             horizontalalignment='center', verticalalignment='bottom', 
             color='white', 
             transform=ax1.get_transform('world'))
    
    # Astropy RGB solution
    ax1.imshow(make_lupton_rgb(imgR, imgG, imgB, 
                               # stretch=0.07, Q=5
                               ), 
               origin='lower', zorder=-1, alpha=1)
    
    # ax1.set_xlim(ax1.get_xlim()[1], ax1.get_xlim()[0]) # flip
    ax1.set_xlabel('RA (deg)', fontsize=15)
    ax1.set_ylabel('DEC (deg)', fontsize=15)
    ax1.minorticks_on()
    ax1.legend(prop={'size': 12}, facecolor='dimgray')
    ax1.set_aspect(1)
    plt.savefig("slit_distribution.jpg", dpi=200, bbox_inches='tight')
    
    return g_cat_return







if __name__ == '__main__':
    # If any, specify slits that were not finished
    slits_not_finished = [15, 24, 45, 63, 109, 112, 116, 122, 129]
    
    # Redshifts
    z_table_filename = "./scripts/redshift_table.xlsx"
    df     = pd.read_excel(z_table_filename, header=None, engine='openpyxl')
    array  = df.to_numpy()
    ztable = array[1:, 0:2] # remove 1st line header, keep 1 remaining
    ltable = array[1:, [0,6,9,12]]
    # bltable = array[1:, [0,14]]
    
    # Add some columns
    g_cat = np.append(array, 
                      np.array([[np.nan, 'slit_RA' ]+[np.nan]*(len(array)-2), # line 15
                                [np.nan, 'slit_DEC']+[np.nan]*(len(array)-2), # line 16
                                [np.nan, 'M_s'     ]+[np.nan]*(len(array)-2), # line 17
                                [np.nan, 'M_s_err' ]+[np.nan]*(len(array)-2), # line 18
                                ]).T, 
                      axis=1)
    
    real_pkl_folder  = './scripts/binospec_pkl/'
    Ms_folder        =  '../bagpipes-KL/'
    spec1d2dfolder02 = '../../RSCH3/UAO-S156-23B-A383/psf/231019/1d2dspecfiles/'
    
    date_of_run = str(input('Please enter the data of run: '))
    # date_of_run = '20260601'
    base_dir = f'../../RSCH3/HPC_database/runs_{date_of_run}/'
    g_cat = posterior_to_cat(g_cat, base_dir, ztable, 
                             slits_not_finished)
    
    # Save shear catalog
    from openpyxl import load_workbook
    new_filename = "summary/redshift_table_with_shear_notedge_constrained.xlsx"
    wb = load_workbook(z_table_filename)
    ws = wb.active # or, wb["Sheet1"]
    n_old_cols, n_rows = ws.max_column, ws.max_row
    new_cols = g_cat[:, n_old_cols:]
    new_cols = np.where(new_cols == 'nan', None, new_cols)
    for i in range(n_rows):
        for j in range(new_cols.shape[1]):
            ws.cell(row=i+1, column=n_old_cols + j + 1,
                    value=new_cols[i, j])
    wb.save(new_filename)
    
    print('Finished reading shear catalog. Plotting...')
    
    img_dir  = '../../RSCH3/HSC_img_A383/'
    sci_fn_R = 'hlsp_clash_subaru_suprimecam_a383_z_2010-2002-v20110405_drz.fits'
    sci_fn_G = 'hlsp_clash_subaru_suprimecam_a383_ip_2005-v20110422_drz.fits'
    # sci_fn_G = 'hlsp_clash_subaru_suprimecam_a383_ic_2002-v20110405_drz.fits'
    sci_fn_B = 'hlsp_clash_subaru_suprimecam_a383_v_2002-2005-2008-v20110405_drz.fits'
    # sci_fn_B = 'hlsp_clash_subaru_suprimecam_a383_rc_2002-2007-2008-v20110405_drz.fits'
    
    img_data_R = read_subaru_img_wcs(img_dir+sci_fn_R, None)
    img_data_G = read_subaru_img_wcs(img_dir+sci_fn_G, None)
    img_data_B = read_subaru_img_wcs(img_dir+sci_fn_B, None)
    imgR, wcsR = img_data_R['science_data'], img_data_R['science_wcs' ]
    imgG, wcsG = img_data_G['science_data'], img_data_G['science_wcs' ]
    imgB, wcsB = img_data_B['science_data'], img_data_B['science_wcs' ]
    
    hdulist01 = fits.open(spec1d2dfolder02+'../obj_abs_slits_extr.fits')
    slitsInfo01 = np.append(hdulist01[4].data, hdulist01[5].data, axis=0)
    maskCenterRA, maskCenterDEC = slitsInfo01['MASK_RA'][0], slitsInfo01['MASK_DEC'][0]
    
    scale = 0.2 # arcsec/pix: 0.2 for HSC image
    imgRc, RAlim, DEClim,       = cutoutimg(
        imgR, wcsR, maskCenterRA, maskCenterDEC, 
        img_width=(25)*60/scale, img_height=(25)*60/scale)
    imgGc, RAlim, DEClim, wcsGc = cutoutimg(
        imgG, wcsG, maskCenterRA, maskCenterDEC, 
        img_width=(25)*60/scale, img_height=(25)*60/scale, outputWCS=True)
    imgBc, RAlim, DEClim,       = cutoutimg(
        imgB, wcsB, maskCenterRA, maskCenterDEC, 
        img_width=(25)*60/scale, img_height=(25)*60/scale)
    
    g_cats = slit_design_plot(g_cat, 
                              maskCenterRA, maskCenterDEC, 
                              imgRc*0.01, imgGc*0.01, imgBc*0.004, wcsGc, 
                              slits_not_finished)
    print('Done.')
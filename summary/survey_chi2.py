import json
import yaml
import numpy as np
import pandas as pd
from pathlib import Path
            
from klm.nautilus_sampler import NautilusSampler
from klm.parameters       import Parameters
# from data_prep.read_save_utils import readinfodat
from binospec_main_fitting_nautilus import make_config_dic, load_mock
from binospec_plot_corner_new  import get_max_num_subdir
# from post_fitting import analyze_percentile

def add_chi2(g_cat, ztable):
    """
        Generate a shear catalog.
    """
    for slit_num in range(1, 143):
        z_spec = ztable[slit_num, 1]
        
        if np.isnan(z_spec)==True:
            print( "\033[43m" + 'WARNING:' + "\033[0m " + 
                  f'Slit {slit_num} skipped because no secure redshift found.\n')
            continue
    
        elif np.isnan(z_spec)==False:
            try:
                data_info = load_mock(real_pkl_folder, Ms_folder, slit_num)
            except FileNotFoundError: 
                print( "\033[43m" + 'WARNING:' + "\033[0m " + 
                      f'Slit {slit_num} skipped because no pkl found.\n')
                continue
            
            # Find the most latest run
            base_dir = '../../../RSCH3/kl_github/'
            full_run_dir_1 = f'{base_dir}runs_nautilus/Slit_{slit_num:03d}/'
            if Path(full_run_dir_1).exists() is False:
                print(f'Slit {slit_num} does not exist.')
                continue
            
            # Load YAML file for config
            with open(fiducial_yaml, "r", encoding="utf-8") as file1:
                fid_params     = yaml.safe_load(file1)
            with open(fitting_yaml, "r", encoding="utf-8") as file2:
                fitting_params = yaml.safe_load(file2)
            
            linespecies = []
            for spec in data_info['spec']:
                linespecies.append(spec['par_meta']['line_species'])
            
            config_dic = make_config_dic(
                linespecies, fitting_params, fid_params, 
                log10_Mstar=data_info['galaxy']['log10_Mstar'], 
                log10_Mstar_err=data_info['galaxy']['log10_Mstar_err'],
                use_line_profile = None, #  'meta' 
                )
            
            date_of_run1 = 'runs_20260328' + '/'
            # date_of_run1 = get_max_num_subdir(full_run_dir_1) + '/'
            best_path1 = f'{full_run_dir_1}{date_of_run1}best_fit.json'
            if Path(best_path1).exists() is False:
                print(f'Slit {slit_num} does not exist.')
                continue
            with open(best_path1, "r") as f:
                best_par = json.load(f)['maximum_likelihood']['point']
            
            inference = NautilusSampler(data_info, config_dic)
            inference.params.params['shared_params'].update({
                **inference.params.params['shared_params'], 
                **Parameters._unflatten(best_par)['shared_params']
                })
            
            img_obs  = np.flip(inference.data_image, axis=1)
            imgDOF   = img_obs.shape[0] * img_obs.shape[1] - len(best_par)
            img_chi2 = inference.calc_image_loglike(
                inference.params.params
                )
            chi2_image_reduced = img_chi2 / imgDOF
            
            chi2_spec_sum, chi2_spec_reduced_avg = 0, 0
            for i in range(len(inference.data_spec)):
                inference.spec_model[i]._init_observable(
                    data_info['galaxy'], 
                    data_info['spec'][i]['par_meta'])
                
                # Update by set. Assign I01 from I01_spec1/2/3 now
                line = linespecies[i//3]
                best_line_pars = {
                    **inference.params.params[f'{line}_params'], 
                    **Parameters._unflatten(best_par)[f'{line}_params']
                    }
                
                # Use I02, I01 in calculation instead
                set_num = i % 3 + 1
                for k in ['I02', 'I01']: #, 'dx_vel', 'dy_vel', 'bkg_level']:
                    best_line_pars[k] = None
                    if k+f'_spec{set_num}' in best_line_pars.keys():
                        best_line_pars[k] = best_line_pars[k+f'_spec{set_num}']
                
                spec0_obs = inference.data_spec[i]
                spec0_msk = inference.mask_spec[i]
                svar0_obs = inference.var_spec[i]
                scon0_obs = inference.cont_model[i]
                spec0_fit = inference.spec_model[i].get_observable(
                    {**inference.params.params[ 'shared_params'],
                     **best_line_pars }
                    )
                spec0_chi2 = inference._loglike_one_slit(
                    data_spec=spec0_obs, 
                    mask_spec=spec0_msk,
                    var_spec=svar0_obs, 
                    cont_spec=scon0_obs,
                    model_spec=spec0_fit)
                specDOF = spec0_obs.shape[0] * spec0_obs.shape[1] - len(best_par)
                
                chi2_spec_sum += spec0_chi2
                chi2_spec_reduced_avg += (spec0_chi2 / specDOF)
                
            chi2_spec_reduced_avg /= len(inference.data_spec)
            
            chi2_joint = img_chi2 + chi2_spec_sum
            chi2_joint_reduced_avg = (chi2_image_reduced + 
                                      chi2_spec_reduced_avg) / 2
            
            g_cat[slit_num+1, -6:] = np.array([[
                img_chi2, chi2_image_reduced, 
                chi2_spec_sum, chi2_spec_reduced_avg,
                chi2_joint, chi2_joint_reduced_avg
                ]])
            
            print(f'INFO:    Slit {slit_num} recorded. 👍')
    
    return g_cat


real_pkl_folder  = 'binospec_pkl/'
shear_table_fnme = "./redshift_table_with_shear.xlsx"
# shear_table_fnme = "./redshift_table.xlsx"
Ms_folder        =  '../../bagpipes-KL/'
spec1d2dfolder02 = '../../../RSCH3/UAO-S156-23B-A383/psf/231019/1d2dspecfiles/'
fiducial_yaml    =  "./config/binospec_fid_params.yaml"
fitting_yaml     =  "./config/binospec_fitting_params.yaml"

if __name__ == '__main__':
    df     = pd.read_excel(shear_table_fnme, header=None, engine='openpyxl')
    array  = df.to_numpy()
    ztable = (df.iloc[2:, [5, 8, 11]]
                .apply(pd.to_numeric, errors='coerce') 
                .mean(axis=1, skipna=True)
                ).to_numpy()
    array[2:, 1] = ztable.T
    
    # Add some columns
    columns = np.array(
        [['chi2', 'chi2_image'            ]+[np.nan]*(len(array)-2),
         [np.nan, 'chi2_image_reduced'    ]+[np.nan]*(len(array)-2),
         [np.nan, 'chi2_spec_sum'         ]+[np.nan]*(len(array)-2),
         [np.nan, 'chi2_spec_reduced_avg' ]+[np.nan]*(len(array)-2),
         [np.nan, 'chi2_joint'            ]+[np.nan]*(len(array)-2),
         [np.nan, 'chi2_joint_reduced_avg']+[np.nan]*(len(array)-2),
         ])
    g_cat = np.append(array, columns.T, axis=1)
    
    g_cat = add_chi2(g_cat, array[1:, 0:2])
    
    # Save shear catalog
    from openpyxl import load_workbook
    new_filename = "./redshift_table_with_shear_chi2.xlsx"
    wb = load_workbook(shear_table_fnme)
    ws = wb.active # or, wb["Sheet1"]
    n_old_cols, n_rows = ws.max_column, ws.max_row
    new_cols = g_cat[:, n_old_cols:]
    new_cols = np.where(new_cols == 'nan', None, new_cols)
    for i in range(n_rows):
        for j in range(new_cols.shape[1]):
            ws.cell(row=i+1, column=n_old_cols + j + 1,
                    value=new_cols[i, j])
    wb.save(new_filename)